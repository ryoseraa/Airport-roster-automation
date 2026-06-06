"""
CNS DUTY ROSTER GENERATOR — CCS Airport Lucknow
================================================
Reads 4 Excel files:
  SHIFT.xlsx          — shift duty employees (name, unit, designation)
  GEN-EMP.xlsx        — general duty employees
  leave.xlsx          — employees on leave (NAME, UNIT, DESIGNATION, TYPE, START, END)
  ADDITIONALDUTY.xlsx — employees who did additional duty

SHIFT CYCLE (per employee, repeats every 4 days):
  Day 1 → B   : 07:00–13:00  Morning duty (solo)
  Day 2 → C   : 13:00–19:00  Afternoon duty (solo)
  Day 3 → DA  : 19:00–07:00  Night duty D+A combined (same person does both)
  Day 4 → OFF : Rest day (mandatory after DA night)
  Day 5 → B   : cycle repeats

  4 employees per unit cover 1 day: one on B, one on C, one on DA, one on OFF.
  Unit suffixes: AMSS=none, EQ.ROOM=e, NAVAIDS=n, ASMGCS/AUTOMATION=a, RADAR=r

GENERAL: Mon–Fri "G" (09:30–18:00), Sat–Sun "OFF"
ADDITIONAL DUTY: 4+ extra hours → Comp Leave due within 3 months
"""

import os
from datetime import date, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Cycle anchor — phase is calculated relative to this date ───
# Phase 0 on REFERENCE_DATE means that employee does B on that day.
REFERENCE_DATE = date(2026, 3, 29)

# CORRECT 4-day phase mapping:
#   0 → B   (Morning   07:00–13:00)
#   1 → C   (Afternoon 13:00–19:00)
#   2 → DA  (Night     19:00–07:00, D and A done together by same person)
#   3 → OFF (Rest — mandatory after DA)

UNIT_SUFFIX = {
    "AMSS":       "",
    "EQ.ROOM":    "e",
    "NAVAIDS":    "n",
    "ASMGCS":     "a",
    "AUTOMATION": "a",
    "RADAR":      "r",
}

COLOUR = {
    "B":   "C6EFCE",   # green  — morning
    "C":   "FFEB9C",   # yellow — afternoon
    "DA":  "9DC3E6",   # blue   — night DA combined
    "OFF": "D9D9D9",   # grey
    "L":   "FF9999",   # red    — leave
    "T":   "E2EFDA",   # light  — training
    "G":   "DDEBF7",   # sky    — general
    "AD":  "FFD966",   # gold   — additional duty
}

def cell_colour(code):
    code = str(code).upper().strip().replace("*", "")
    if code == "OFF":             return COLOUR["OFF"]
    if code in ("L", "LEAVE"):   return COLOUR["L"]
    if code == "G":               return COLOUR["G"]
    if code == "T":               return COLOUR["T"]
    if code.startswith("B"):      return COLOUR["B"]
    if code.startswith("C"):      return COLOUR["DA"]   # C*D* night combined
    return "FFFFFF"

def tborder():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

# ── Read all 4 input files ──────────────────────────────────────
def read_inputs(shift_path, gen_path, leave_path, addl_path):
    def clean(v):
        return str(v).strip().replace("\t","") if v and str(v).strip() not in ("nan","None") else ""

    # SHIFT employees
    wb = load_workbook(shift_path, data_only=True)
    ws = wb.active
    shift_emps = []
    unit_pools  = {}   # unit -> list of employees in order
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = clean(row[0]); unit = clean(row[1]); desig = clean(row[2])
        if not name: continue
        unit = unit.upper().strip()
        emp = {"name": name, "unit": unit, "designation": desig,
               "shift_phase": 0, "additional_duty": False}
        shift_emps.append(emp)
        unit_pools.setdefault(unit, []).append(emp)

    # Assign phases 0,1,2,3 to each person in a unit pool
    for unit, pool in unit_pools.items():
        for i, emp in enumerate(pool):
            emp["shift_phase"] = i % 4

    # GENERAL employees
    wb2 = load_workbook(gen_path, data_only=True)
    ws2 = wb2.active
    gen_emps = []
    for row in ws2.iter_rows(min_row=2, values_only=True):
        name = clean(row[0]); unit = clean(row[1]); desig = clean(row[2])
        if not name: continue
        gen_emps.append({"name": name, "unit": unit.upper().strip(),
                         "designation": desig})

    # LEAVE — now supports start/end dates (cols: NAME, UNIT, DESIGNATION, TYPE, START, END)
    wb3 = load_workbook(leave_path, data_only=True)
    ws3 = wb3.active
    leave_records = []   # list of {name, type, start, end}
    for row in ws3.iter_rows(min_row=2, values_only=True):
        name  = clean(row[0])
        ltype = clean(row[3]).upper() if row[3] else "GENERAL"
        if not name: continue
        # Parse start date (col 4)
        try:
            start = row[4].date() if hasattr(row[4], "date") else date.fromisoformat(str(row[4])[:10])
        except Exception:
            start = None
        # Parse end date (col 5)
        try:
            end = row[5].date() if hasattr(row[5], "date") else date.fromisoformat(str(row[5])[:10])
        except Exception:
            end = start   # if no end date, treat as single day
        if ltype not in ("SHIFT","GENERAL"):
            ltype = "GENERAL"
        leave_records.append({
            "name":  name.upper(),
            "type":  ltype,
            "start": start,
            "end":   end,
        })
    # keep backward compat: also expose as on_leave sets (used for count display)
    on_leave = {
        "SHIFT":   {r["name"] for r in leave_records if r["type"] == "SHIFT"},
        "GENERAL": {r["name"] for r in leave_records if r["type"] == "GENERAL"},
    }

    # ADDITIONAL DUTY
    wb4 = load_workbook(addl_path, data_only=True)
    ws4 = wb4.active
    addl_names = set()
    for row in ws4.iter_rows(min_row=2, values_only=True):
        name = clean(row[0])
        if name: addl_names.add(name.upper())

    # Mark additional duty on shift employees
    for emp in shift_emps:
        if emp["name"].upper() in addl_names:
            emp["additional_duty"] = True

    return shift_emps, gen_emps, on_leave, addl_names, leave_records

# ── Shift code for a given employee + date ──────────────────────
def get_shift_code(emp, target_date):
    """
    Correct 4-day cycle:
      Phase 0 → B    07:00-13:00  Morning (solo)
      Phase 1 → C    13:00-19:00  Afternoon (solo)
      Phase 2 → DA   19:00-07:00  Night D+A combined (same person)
      Phase 3 → OFF  Rest day (mandatory after DA night)

    Example for employee starting at phase 0:
      1 Apr (Mon) → B
      2 Apr (Tue) → C
      3 Apr (Wed) → DA  (works 7pm to 7am next day)
      4 Apr (Thu) → OFF (rest after night)
      5 Apr (Fri) → B   (cycle repeats)
    """
    days   = (target_date - REFERENCE_DATE).days
    phase  = (emp["shift_phase"] + days) % 4
    suffix = UNIT_SUFFIX.get(emp["unit"], "")

    if phase == 0:
        return "B" + suffix                                          # Morning
    elif phase == 1:
        return "C" + suffix                                          # Afternoon
    elif phase == 2:
        # Night: D then A done together by same person
        return ("C" + suffix + "D" + suffix) if suffix else "DA"   # e.g. CaDa, CnDn
    else:  # phase == 3
        return "OFF"                                                  # Rest after night

# ── Generate schedules ──────────────────────────────────────────

def is_on_leave_date(name, target_date, leave_records, ltype):
    """Returns True if employee is on leave on target_date (checks date range)."""
    for rec in leave_records:
        if rec["name"] == name.upper() and rec["type"] == ltype:
            if rec["start"] and rec["end"]:
                if rec["start"] <= target_date <= rec["end"]:
                    return True
    return False

def generate_shift_schedule(shift_emps, on_leave, week_start, leave_records):
    schedule = {}
    for emp in shift_emps:
        weekly = []
        for offset in range(7):
            day = week_start + timedelta(days=offset)
            if is_on_leave_date(emp["name"], day, leave_records, "SHIFT"):
                code = "L"
            else:
                code = get_shift_code(emp, day)
                if emp["additional_duty"] and code not in ("OFF","L"):
                    code = code + "*"   # * marks additional duty
            weekly.append(code)
        schedule[emp["name"]] = {
            "designation": emp["designation"],
            "unit":        emp["unit"],
            "codes":       weekly,
            "addl":        emp["additional_duty"],
        }
    return schedule

def generate_general_schedule(gen_emps, on_leave, week_start, leave_records):
    schedule = {}
    for emp in gen_emps:
        weekly = []
        for offset in range(7):
            day = week_start + timedelta(days=offset)
            if is_on_leave_date(emp["name"], day, leave_records, "GENERAL"):
                code = "L"
            elif day.weekday() < 5:   # Mon–Fri
                code = "G"
            else:
                code = "OFF"
            weekly.append(code)
        schedule[emp["name"]] = {
            "designation": emp["designation"],
            "unit":        emp["unit"],
            "codes":       weekly,
        }
    return schedule

# ── Coverage check ──────────────────────────────────────────────
def check_coverage(shift_schedule, week_start):
    """
    Every unit must have exactly one person on each of:
      B  (Morning   07:00-13:00)
      C  (Afternoon 13:00-19:00)
      DA (Night     19:00-07:00 combined)
    per day. The 4th person is always on OFF.
    """
    by_unit = {}
    for name, data in shift_schedule.items():
        by_unit.setdefault(data["unit"], []).append(data["codes"])

    alerts = []
    for unit, all_codes in by_unit.items():
        suffix = UNIT_SUFFIX.get(unit, "")
        needed = {
            "Morning B  (07:00-13:00)":   "B" + suffix,
            "Afternoon C (13:00-19:00)":  "C" + suffix,
            "Night DA   (19:00-07:00)":   ("C"+suffix+"D"+suffix) if suffix else "CDA",
        }
        for day_idx in range(7):
            day       = week_start + timedelta(days=day_idx)
            day_codes = [str(c).replace("*", "") for c in [codes[day_idx] for codes in all_codes]]
            for slot, slot_code in needed.items():
                if not any(slot_code in c for c in day_codes):
                    alerts.append(
                        f"UNCOVERED — {unit} | {slot} | {day.strftime('%a %d %b')}"
                    )
    return alerts

# ── Excel Export ────────────────────────────────────────────────
def write_sheet(ws, schedule, week_start, hdr_colour, title1, title2):
    days     = [week_start + timedelta(days=i) for i in range(7)]
    day_names = ["SUN","MON","TUE","WED","THU","FRI","SAT"]

    # Title rows
    ws.merge_cells("A1:L1")
    ws["A1"] = title1
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:L2")
    ws["A2"] = title2
    ws["A2"].font = Font(bold=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Column headers — row 3
    hfill = PatternFill("solid", fgColor=hdr_colour)
    hfont = Font(bold=True, color="FFFFFF", size=10)
    headers = ["#", "NAME", "DESIGNATION", "UNIT"] + \
              [f"{day_names[d.weekday()]}\n{d.strftime('%d/%m')}" for d in days] + \
              ["REMARKS"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = tborder()
    ws.row_dimensions[3].height = 30

    row_num   = 4
    serial    = 1
    prev_unit = None

    for name, data in schedule.items():
        # Unit section divider
        if data["unit"] != prev_unit:
            ws.merge_cells(start_row=row_num, start_column=1,
                           end_row=row_num, end_column=12)
            c = ws.cell(row=row_num, column=1)
            c.value = f"  ◆  {data['unit']}"
            c.font  = Font(bold=True, size=10, color="1F3864")
            c.fill  = PatternFill("solid", fgColor="BDD7EE")
            c.alignment = Alignment(vertical="center")
            row_num  += 1
            prev_unit = data["unit"]

        ws.cell(row=row_num, column=1).value = serial
        ws.cell(row=row_num, column=2).value = name
        ws.cell(row=row_num, column=3).value = data["designation"]
        ws.cell(row=row_num, column=4).value = data["unit"]

        remarks = []
        if data.get("addl"):
            remarks.append("Additional Duty — Comp Leave Due")
        ws.cell(row=row_num, column=12).value = " | ".join(remarks)
        ws.cell(row=row_num, column=12).font  = Font(size=8, color="C00000")

        for di, code in enumerate(data["codes"]):
            col  = 5 + di
            c    = ws.cell(row=row_num, column=col)
            c.value     = str(code)
            c.fill      = PatternFill("solid", fgColor=cell_colour(str(code)))
            c.font      = Font(bold=True, size=9)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = tborder()

        for col in [1, 2, 3, 4, 12]:
            c = ws.cell(row=row_num, column=col)
            c.border    = tborder()
            c.font      = Font(size=8, color="C00000") if col == 12 else Font(size=9)
            c.alignment = Alignment(vertical="center", wrap_text=(col in [2,3,4]))

        ws.row_dimensions[row_num].height = 17
        row_num += 1
        serial  += 1

    return row_num

def add_footer(ws, start_row, is_shift=True):
    # Legend
    r = start_row + 1
    ws.cell(row=r, column=1).value = "LEGEND:"
    ws.cell(row=r, column=1).font  = Font(bold=True, size=9)

    if is_shift:
        items = [
            ("B",   "07:00–13:00  Morning duty (solo)"),
            ("C",   "13:00–19:00  Afternoon duty (solo)"),
            ("CDA", "19:00–07:00  Night duty D+A combined (same person)"),
            ("OFF", "Rest Day (mandatory after DA night)"),
            ("L",   "Leave"),
            ("*",   "Additional Duty — Comp Leave due within 3 months"),
        ]
    else:
        items = [
            ("G","09:30–18:00  General Duty (Mon–Fri)"),
            ("OFF","Rest Day (Sat–Sun)"),
            ("L","Leave"),
        ]

    for i, (code, desc) in enumerate(items):
        c = ws.cell(row=r+1+i, column=1)
        c.value = f"  {code}  =  {desc}"
        c.fill  = PatternFill("solid", fgColor=cell_colour(code))
        c.font  = Font(size=8)
        ws.merge_cells(start_row=r+1+i, start_column=1,
                       end_row=r+1+i, end_column=6)

    notes_row = r + len(items) + 3
    notes = [
        "NOTES:",
        "1. Night shift CD (19:00–07:00) is done by one person continuously. Next day is OFF.",
        "2. If Radar/Automation shift is uncovered, both units are clubbed.",
        "3. If AMSS Booking or AMSS Supervisor is absent, both duties are clubbed.",
        "4. SSO (Shift I/C) = most senior officer present in the shift.",
        "5. Additional Duty (*): Employee worked 4+ extra hours → entitled to 1 Comp Leave (expires in 3 months).",
        "6. No duty is changed without the approval of undersigned.",
    ]
    for i, note in enumerate(notes):
        ws.merge_cells(start_row=notes_row+i, start_column=1,
                       end_row=notes_row+i, end_column=12)
        c = ws.cell(row=notes_row+i, column=1, value=note)
        c.font = Font(bold=(i==0), size=8, italic=(i>0))

    sig_row = notes_row + len(notes) + 2
    ws.cell(row=sig_row, column=9).value = "Approved by:"
    ws.cell(row=sig_row, column=10).value = "कृते सी.एन.एस. प्रभारी"
    ws.cell(row=sig_row, column=10).font  = Font(bold=True, size=10)
    ws.cell(row=sig_row+1, column=10).value = "सीसीएसआई, एअरपोर्ट, लखनऊ"

def set_widths(ws):
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14
    for i in range(5, 12):
        ws.column_dimensions[get_column_letter(i)].width = 11
    ws.column_dimensions["L"].width = 28

# ── MAIN ────────────────────────────────────────────────────────
def main():
    print("="*58)
    print("  CNS DUTY ROSTER GENERATOR — CCS Airport, Lucknow")
    print("="*58)
    print("\nSHIFT CYCLE: B (07-13) → C (13-19) → DA (19-07) → OFF")
    print("Each unit needs 4 employees (one per phase).\n")

    shift_path = "SHIFT.xlsx"
    gen_path   = "GEN-EMP.xlsx"
    leave_path = "leave.xlsx"
    addl_path  = "ADDITIONALDUTY.xlsx"

    for f in [shift_path, gen_path, leave_path, addl_path]:
        if not os.path.exists(f):
            print(f"ERROR: Missing file — {f}")
            return

    print("Reading your 4 Excel files...")
    shift_emps, gen_emps, on_leave, addl_names, leave_records = read_inputs(
        shift_path, gen_path, leave_path, addl_path)

    print(f"  Shift employees  : {len(shift_emps)}")
    print(f"  General employees: {len(gen_emps)}")
    print(f"  On leave (shift) : {len(on_leave['SHIFT'])}")
    print(f"  On leave (gen)   : {len(on_leave['GENERAL'])}")
    print(f"  Additional duty  : {len(addl_names)}")

    # ── Date input ──────────────────────────────────────────────
    print("\n" + "─"*58)
    print("ROSTER DATE RANGE")
    print("─"*58)
    print("Enter the START date of the week you want to generate.")
    print("Format: YYYY-MM-DD   Example: 2026-06-08")
    print("(The roster will cover 7 days from the date you enter.)\n")

    while True:
        week_input = input("Week start date: ").strip()
        try:
            week_start = date.fromisoformat(week_input)
            break
        except ValueError:
            print("  Invalid format. Please use YYYY-MM-DD (e.g. 2026-06-08)")

    week_end = week_start + timedelta(days=6)
    print(f"\nGenerating roster for: "
          f"{week_start.strftime('%d %b %Y')} → {week_end.strftime('%d %b %Y')}")

    # ── Generate ─────────────────────────────────────────────────
    shift_sched   = generate_shift_schedule(shift_emps, on_leave, week_start, leave_records)
    general_sched = generate_general_schedule(gen_emps, on_leave, week_start, leave_records)
    alerts        = check_coverage(shift_sched, week_start)

    if alerts:
        print(f"\n{'─'*58}\n⚠  COVERAGE ALERTS:")
        for a in alerts:
            print("   •", a)
        print('─'*58)
    else:
        print("✓ All shift slots covered for this week.")

    # ── Export ───────────────────────────────────────────────────
    os.makedirs("output", exist_ok=True)
    out = os.path.join("output", f"CNS_Roster_{week_start.strftime('%Y-%m-%d')}.xlsx")

    days = [week_start + timedelta(days=i) for i in range(7)]
    dr   = f"{week_start.strftime('%d %b')} to {days[-1].strftime('%d %b %Y')}"

    wb  = Workbook()
    ws1 = wb.active
    ws1.title = "Shift Duty"
    end1 = write_sheet(ws1, shift_sched, week_start, "2F5496",
        "Airports Authority of India — CCSI Airport, Lucknow",
        f"CNS Executive Shift Duty Roster  ·  {dr}")
    add_footer(ws1, end1, is_shift=True)
    set_widths(ws1)

    ws2 = wb.create_sheet("General Duty")
    end2 = write_sheet(ws2, general_sched, week_start, "375623",
        "Airports Authority of India — CCSI Airport, Lucknow",
        f"CNS Executive General Duty Roster  ·  {dr}")
    add_footer(ws2, end2, is_shift=False)
    set_widths(ws2)

    wb.save(out)
    print(f"\n✓ Roster saved → {os.path.abspath(out)}")

    # ── Ask if another week needed ───────────────────────────────
    print("\n" + "─"*58)
    again = input("Generate another week? (y/n): ").strip().lower()
    if again == "y":
        main()

if __name__ == "__main__":
    main()
